
'''
import xlrd
 

loc = 'C:\python prac sec\Java-WeCP_review.xls'
 
# To open Workbook
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
 
# For row 0 and column 0
print(sheet.cell_value(0, 0))


'''


from openpyxl import * #load_workbook

workbook = load_workbook(filename="C:\python prac sec\Java - WeCP review.xlsx")

#sheet = workbook.active
#print(workbook.worksheets)
#for i in workbook.worksheets:
        #sheet=i
        #print(i)

sheet=workbook.worksheets[3]
#f= open("Excel_data_1.txt","w+")

for i in sheet.iter_rows(min_row=2,max_row=46,
                         min_col=5,max_col=11,
                         values_only=True):
        for j in i:
                if j!=None:
                        print(j)
                        #f.write(j)
                #print('\n')
        print('\n','#'*40,'\n')
        #s='#'*40
        #f.write(s)
#f.close()
print('Done')
